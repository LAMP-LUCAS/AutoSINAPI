"""
Utilitários centralizados para o sistema SINAPI
Este módulo contém todas as funções e classes comuns utilizadas pelos outros módulos do sistema.
"""
import logging
import unicodedata
import pandas as pd
import os
import zipfile
import requests
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Any, Union
from pathlib import Path
from sqlalchemy import create_engine, text
from tqdm import tqdm
from openpyxl import load_workbook

class SinapiLogger:
    """Classe para gerenciar logs do sistema SINAPI"""
    
    def __init__(self, nome: str, level: str = 'info'):
        """
        Inicializa o logger com nome e nível específicos
        Args:
            nome (str): Nome do logger
            level (str): Nível de log ('debug', 'info', 'warning', 'error', 'critical')
        """
        self.logger = logging.getLogger(nome)
        self.configure(level)
    
    def configure(self, level: str = 'info') -> None:
        """Configura o logger com o nível especificado"""
        levels = {
            'debug': logging.DEBUG,
            'info': logging.INFO,
            'warning': logging.WARNING,
            'error': logging.ERROR,
            'critical': logging.CRITICAL
        }
        
        self.logger.setLevel(levels.get(level.lower(), logging.INFO))
        
        if self.logger.hasHandlers():
            self.logger.handlers.clear()
            
        handler = logging.StreamHandler()
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        handler.setFormatter(formatter)
        self.logger.addHandler(handler)
    
    def log(self, level: str, message: str) -> None:
        """Registra uma mensagem no nível especificado"""
        getattr(self.logger, level.lower())(message)

class ExcelProcessor:
    """Classe para processamento de arquivos Excel do SINAPI"""
    
    def __init__(self):
        self.logger = SinapiLogger("ExcelProcessor")
    
    def scan_directory(self, diretorio: str = None, formato: str = 'xlsx', data: bool = False) -> Dict:
        """
        Escaneia um diretório em busca de arquivos Excel
        Args:
            diretorio (str): Caminho do diretório
            formato (str): Formato dos arquivos ('xlsx', 'xls', etc)
            data (bool): Se True, processa os dados das planilhas
        Returns:
            Dict: Resultados do processamento
        """
        if not diretorio:
            diretorio = os.getcwd()
        diretorio = Path(diretorio).resolve()
        
        self.logger.log('info', f'Escaneando o diretório: {diretorio}')
        resultado = {}
        
        try:
            for arquivo in os.listdir(diretorio):
                if not arquivo.lower().endswith(formato.lower()):
                    continue
                    
                caminho = diretorio / arquivo
                self.logger.log('info', f'Processando: {arquivo}')
                
                if data:
                    try:
                        wb = load_workbook(caminho, read_only=True)
                        planilhas_info = []
                        for nome_planilha in wb.sheetnames:
                            ws = wb[nome_planilha]
                            dados = self.get_sheet_data(ws)
                            planilhas_info.append((nome_planilha, dados))
                        resultado[arquivo] = planilhas_info
                        wb.close()
                    except Exception as e:
                        self.logger.log('error', f"Erro ao processar {arquivo}: {str(e)}")
                else:
                    resultado[arquivo] = str(caminho)
                    
        except Exception as e:
            self.logger.log('error', f"Erro ao escanear diretório: {str(e)}")
            
        self.logger.log('info', f'Encontrados {len(resultado)} arquivos')
        return resultado

    def get_sheet_data(self, ws) -> List[int]:
        """
        Extrai informações básicas de uma planilha Excel
        Args:
            ws: Worksheet do openpyxl
        Returns:
            List[int]: [total_cells, n_rows, n_cols]
        """
        if not any(ws.cell(row=1, column=col).value for col in [1,2]) and not ws.cell(row=2, column=1).value:
            return [0, 0, 0]
        
        total_cells = 0
        for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row,
                              min_col=ws.min_column, max_col=ws.max_column,
                              values_only=True):
            total_cells += sum(1 for cell in row if cell is not None and (not isinstance(cell, str) or cell.strip()))
        
        return [total_cells, ws.max_row - ws.min_row + 1, ws.max_column - ws.min_column + 1]

class FileManager:
    """Classe para gerenciamento de arquivos do SINAPI"""
    
    def __init__(self):
        self.logger = SinapiLogger("FileManager")
    
    def normalize_text(self, text: str) -> str:
        """
        Normaliza um texto removendo acentos e convertendo para maiúsculo
        Args:
            text (str): Texto a ser normalizado
        Returns:
            str: Texto normalizado
        """
        if not text:
            return ''
        
        text = str(text).strip().upper()
        text = text.replace('\n', ' ')
        text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('utf-8')
        return text.replace(' ', '_').replace('-', '_').replace('  ', ' ').strip()

    def normalize_files(self, path: Union[str, Path], extension: str = 'xlsx') -> List[str]:
        """
        Normaliza nomes de arquivos em um diretório
        Args:
            path: Caminho do diretório
            extension (str): Extensão dos arquivos
        Returns:
            List[str]: Lista de nomes normalizados
        """
        path = Path(path)
        extension = extension.strip().lower().lstrip('*.')
        normalized_names = []
        
        for file in path.glob(f'*.{extension}'):
            new_name = self.normalize_text(file.name)
            new_path = file.parent / new_name
            
            if file != new_path:
                try:
                    file.rename(new_path)
                    self.logger.log('info', f'Arquivo renomeado: {file} -> {new_path}')
                except Exception as e:
                    self.logger.log('error', f'Erro ao renomear {file}: {str(e)}')
                    
            normalized_names.append(new_name)
            
        return normalized_names

class SinapiDownloader:
    """Classe para download de arquivos do SINAPI"""
    
    def __init__(self, cache_minutes: int = 10):
        self.logger = SinapiLogger("SinapiDownloader")
        self.cache_minutes = cache_minutes
        self.log_file = "sinap_webscraping_download_log.json"
    
    def download_file(self, ano: str, mes: str, formato: str = 'xlsx') -> Optional[str]:
        """
        Baixa arquivo do SINAPI se necessário
        Args:
            ano (str): Ano de referência (YYYY)
            mes (str): Mês de referência (MM)
            formato (str): Formato do arquivo ('xlsx' ou 'pdf')
        Returns:
            Optional[str]: Caminho do arquivo baixado ou None se falhou
        """
        if not self._validar_parametros(ano, mes, formato):
            return None
            
        if not self._pode_baixar(ano, mes):
            return None
            
        url = f'https://www.caixa.gov.br/Downloads/sinapi-relatorios-mensais/SINAPI-{ano}-{mes}-formato-{formato}.zip'
        folder_name = f'{ano}_{mes}'
        zip_path = Path(folder_name) / f'SINAPI-{ano}-{mes}-formato-{formato}.zip'
        
        if zip_path.exists():
            self.logger.log('info', f'Arquivo já existe: {zip_path}')
            return str(zip_path)
            
        try:
            os.makedirs(folder_name, exist_ok=True)
            self._download_with_retry(url, zip_path)
            self._registrar_download(ano, mes)
            return str(zip_path)
        except Exception as e:
            self.logger.log('error', f'Erro no download: {str(e)}')
            return None

    def unzip_file(self, zip_path: Union[str, Path]) -> Optional[str]:
        """
        Extrai um arquivo ZIP do SINAPI
        Args:
            zip_path: Caminho do arquivo ZIP
        Returns:
            Optional[str]: Caminho da pasta extraída ou None se falhou
        """
        zip_path = Path(zip_path)
        if not zip_path.exists():
            self.logger.log('error', f'Arquivo não existe: {zip_path}')
            return None
            
        extraction_path = zip_path.parent / zip_path.stem
        
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(extraction_path)
            self.logger.log('info', f'Arquivos extraídos em: {extraction_path}')
            return str(extraction_path)
        except Exception as e:
            self.logger.log('error', f'Erro ao extrair: {str(e)}')
            return None

    def _validar_parametros(self, ano: str, mes: str, formato: str) -> bool:
        """Valida os parâmetros de entrada"""
        try:
            if len(ano) != 4 or len(mes) != 2:
                raise ValueError("Ano deve ter 4 dígitos e mês deve ter 2 dígitos")
            if int(mes) < 1 or int(mes) > 12:
                raise ValueError("Mês deve estar entre 01 e 12")
            if formato not in ['xlsx', 'pdf']:
                raise ValueError("Formato deve ser 'xlsx' ou 'pdf'")
            return True
        except Exception as e:
            self.logger.log('error', f'Parâmetros inválidos: {str(e)}')
            return False

    def _pode_baixar(self, ano: str, mes: str) -> bool:
        """Verifica se já passou o tempo mínimo desde o último download"""
        chave = f"{ano}_{mes}"
        agora = datetime.now()
        
        if not os.path.exists(self.log_file):
            return True
            
        try:
            with open(self.log_file, "r") as f:
                log = json.load(f)
            ultimo = log.get(chave)
            if ultimo:
                ultimo_dt = datetime.fromisoformat(ultimo)
                if agora - ultimo_dt < timedelta(minutes=self.cache_minutes):
                    tempo_restante = timedelta(minutes=self.cache_minutes) - (agora - ultimo_dt)
                    self.logger.log('warning', 
                        f"Download recente detectado. Aguarde {tempo_restante} antes de tentar novamente.")
                    return False
        except Exception as e:
            self.logger.log('error', f'Erro ao ler log: {str(e)}')
            
        return True

    def _registrar_download(self, ano: str, mes: str) -> None:
        """Registra a data/hora do download no log"""
        chave = f"{ano}_{mes}"
        agora = datetime.now().isoformat()
        log = {}
        
        if os.path.exists(self.log_file):
            try:
                with open(self.log_file, "r") as f:
                    log = json.load(f)
            except Exception:
                pass
                
        log[chave] = agora
        
        with open(self.log_file, "w") as f:
            json.dump(log, f)

    def _download_with_retry(self, url: str, zip_path: Path, max_retries: int = 3) -> None:
        """Faz o download com retry em caso de falha"""
        session = requests.Session()
        adapter = requests.adapters.HTTPAdapter(max_retries=max_retries)
        session.mount('https://', adapter)
        
        try:
            response = session.get(url, timeout=30, allow_redirects=True)
            response.raise_for_status()
            
            with open(zip_path, 'wb') as f:
                f.write(response.content)
            self.logger.log('info', f'Download concluído: {zip_path}')
            
        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 404:
                self.logger.log('error', 'Arquivo não encontrado no servidor')
            else:
                self.logger.log('error', f'Erro HTTP: {str(e)}')
            raise
            
        except Exception as e:
            self.logger.log('error', f'Erro no download: {str(e)}')
            raise

# Classes existentes mantidas (DatabaseManager, etc)
