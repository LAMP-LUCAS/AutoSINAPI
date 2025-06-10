"""
Utilitários centralizados para o sistema SINAPI
Este módulo contém todas as funções e classes comuns utilizadas pelos outros módulos do sistema.
"""
import logging
import unicodedata
import pandas as pd
import re
import os
import zipfile
import requests
import time
import numpy as np
from time import sleep
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Any, Union
from pathlib import Path
import sqlalchemy
from sqlalchemy import create_engine, text, URL , make_url
from tqdm import tqdm
from openpyxl import load_workbook
import json
import random

class SinapiLogger:
    """Classe para gerenciar logs do sistema SINAPI"""
    
    def __init__(self, nome: str, level: str = 'info'):
        """
        Inicializa o logger com nome e nível específicos
        Args:
            nome (str): Nome do logger
            level (str): Nível de log ('debug', 'info', 'warning', 'error', 'critical')
        """
        self.logger = logging.getLogger(nome)
        self.configure(level)
    
    def configure(self, level: str = 'info') -> None:
        """Configura o logger com o nível especificado"""
        levels = {
            'debug': logging.DEBUG,
            'info': logging.INFO,
            'warning': logging.WARNING,
            'error': logging.ERROR,
            'critical': logging.CRITICAL
        }
        
        self.logger.setLevel(levels.get(level.lower(), logging.INFO))
        
        if self.logger.hasHandlers():
            self.logger.handlers.clear()
            
        handler = logging.StreamHandler()
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        handler.setFormatter(formatter)
        self.logger.addHandler(handler)
    
    def log(self, level: str, message: str) -> None:
        """Registra uma mensagem no nível especificado"""
        getattr(self.logger, level.lower())(message)

class ExcelProcessor:
    """Classe para processamento de arquivos Excel do SINAPI"""
    
    def __init__(self):
        self.logger = SinapiLogger("ExcelProcessor")
    
    def scan_directory(self, diretorio: str = None, formato: str = 'xlsx', data: bool = False,sheet: dict=None,refuse_List: list = None) -> Dict:
        """
        Escaneia um diretório em busca de arquivos Excel
        Args:
            diretorio (str): Caminho do diretório
            formato (str): Formato dos arquivos ('xlsx', 'xls', etc)
            data (bool): Se True, processa os dados das planilhas
            sheet (dict): {sheet_name:sheet_path} Dicionário com as planilhas a terem os dados extraídos.
        Returns:
            Dict: Resultados do processamento
        """
        if not diretorio:
            diretorio = os.getcwd()
        diretorio = Path(diretorio).resolve()
        
        self.logger.log('info', f'Escaneando o diretório: {diretorio}')
        resultado = {}

        if refuse_List is None:
            refuse_List = ['Analítico com Custo','ANALITICO_COM_CUSTO']
        
        try:
            for arquivo in os.listdir(diretorio):
                if not arquivo.lower().endswith(formato.lower()):
                    continue

                caminho = diretorio / arquivo
                self.logger.log('info', f'Verificando: {arquivo}')
                
                if data and sheet:
                    if data and isinstance(sheet, dict) and arquivo in list(sheet.keys()):
                        for sheetName in list(sheet.keys()):
                            if sheetName  in refuse_List:
                                continue
                            else:
                                try:
                                    self.logger.log('info', f'      Processando: {arquivo}')
                                    wb = load_workbook(caminho, read_only=True)
                                    planilhas_info = []
                                    for nome_planilha in wb.sheetnames:
                                        if nome_planilha in refuse_List:
                                            continue
                                        else:
                                            ws = wb[nome_planilha]
                                            dados = self.get_sheet_data(ws)
                                            planilhas_info.append((nome_planilha, dados))
                                    resultado[arquivo] = planilhas_info
                                    wb.close()
                                
                                except Exception as e:
                                    self.logger.log('error', f'Erro ao processar {arquivo} no caminho "{sheet}" : {str(e)}\n   {list(sheet.keys())}')
                    
                elif data and not sheet:
                        try:
                            self.logger.log('info', f'      Processando {arquivo}')
                            wb = load_workbook(caminho, read_only=True)
                            planilhas_info = []
                            for nome_planilha in wb.sheetnames:
                                if nome_planilha in refuse_List:
                                    continue
                                else:
                                    ws = wb[nome_planilha]
                                    dados = self.get_sheet_data(ws)
                                    planilhas_info.append((nome_planilha, dados))
                            resultado[arquivo] = planilhas_info
                            wb.close()
                        except Exception as e:
                            self.logger.log('error', f"Erro ao processar {arquivo}: {str(e)}")
                
                else:
                    self.logger.log('info', f'Coletando nome e caminho do arquivo: {arquivo}')
                    resultado[arquivo] = str(caminho)
        
        except Exception as e:
            self.logger.log('error', f"Erro ao escanear diretório: {str(e)}")
            
        self.logger.log('info', f'Encontrados {len(resultado)} arquivos')
        return resultado

    def get_sheet_data(self, ws) -> List[int]:
        """
        Extrai informações básicas de uma planilha Excel
        Args:
            ws: Worksheet do openpyxl
        Returns:
            List[int]: [total_cells, n_rows, n_cols]
        """
        if not any(ws.cell(row=1, column=col).value for col in [1,2]) and not ws.cell(row=2, column=1).value:
            return [0, 0, 0]
        
        total_cells = 0
        for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row,
                              min_col=ws.min_column, max_col=ws.max_column,
                              values_only=True):
            total_cells += sum(1 for cell in row if cell is not None and (not isinstance(cell, str) or cell.strip()))
        
        return [total_cells, ws.max_row - ws.min_row + 1, ws.max_column - ws.min_column + 1]

class FileManager:
    """Classe para gerenciamento de arquivos do SINAPI"""
    
    def __init__(self):
        self.logger = SinapiLogger("FileManager")
    
    def normalize_text(self, text: str) -> str:
        """
        Normaliza um texto removendo acentos e convertendo para maiúsculo
        Args:
            text (str): Texto a ser normalizado
        Returns:
            str: Texto normalizado
        """
        if not text:
            return ''
        
        text = str(text).strip().upper()
        text = text.replace('\n', ' ')
        text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('utf-8')
        return text.replace(' ', '_').replace('-', '_').replace('  ', ' ').strip()

    def normalize_files(self, path: Union[str, Path], extension: str = 'xlsx') -> List[str]:
        """
        Normaliza nomes de arquivos em um diretório
        Args:
            path: Caminho do diretório
            extension (str): Extensão dos arquivos
        Returns:
            List[str]: Lista de nomes normalizados
        """
        path = Path(path)
        extension = extension.strip().lower().lstrip('*.')
        normalized_names = []
        
        for file in path.glob(f'*.{extension}'):
            self.logger.log('debug', f'Avaliando arquivo: {file}')
            new_name = self.normalize_text(file.stem) + '.' + extension
            new_path = file.parent / new_name
            self.logger.log('debug', f'Novo nome para {file.name} normalizado: {new_name}')
            
            if file.resolve() != new_path.resolve():
                self.logger.log('debug', f'Caminho de origem e destino são diferentes. Tentando renomear.')
                try:
                    file.rename(new_path)
                    self.logger.log('info', f'Arquivo renomeado: {file.name} -> {new_path.name}')
                    normalized_names.append(new_name)
                except Exception as e:
                    self.logger.log('error', f'Erro ao renomear {file}: {str(e)}')

                except:
                    self.logger.log('warning', f'Não foi possível renomear {file.name} para {new_name} pois um arquivo com este nome já existe.')
                    normalized_names.append(new_path.name)
                
            else:
                self.logger.log('debug', f'Arquivo {file.name} já está normalizado.')
                normalized_names.append(file.name)

        self.logger.log('info', f'Nomes normalizados: {str(normalized_names).replace("[","").replace("]","").replace("\'","")}')    
        return normalized_names

    def read_sql_secrets(self, secrets_path: str) -> tuple:
        """Lê credenciais SQL com logging"""
        try:
            with open(secrets_path, 'r') as f:
                content = f.read()
            
            credentials = parse_secrets(content)
            required_keys = {'DB_USER', 'DB_PASSWORD', 'DB_HOST', 'DB_PORT', 'DB_NAME'}
            
            if not required_keys.issubset(credentials):
                missing = required_keys - set(credentials.keys())
                raise ValueError(f"Credenciais incompletas. Faltando: {', '.join(missing)}")
            
            return (
                credentials['DB_USER'],
                credentials['DB_PASSWORD'],
                credentials['DB_HOST'],
                credentials['DB_PORT'],
                credentials['DB_NAME'],
                credentials.get('DB_INITIAL_DB', 'postgres')
            )
        except Exception as e:
            self.logger.log('error', f"Erro ao ler secrets: {e}")
            raise

class DatabaseManager:
    """Classe para gerenciar operações de banco de dados"""
    
    def __init__(self, connection_string: str, log_level: str = 'info'):
        """
        Inicializa o gerenciador de banco de dados
        Args:
            connection_string: String de conexão SQLAlchemy
            log_level: Nível de log
        """
        self.engine = create_engine(connection_string)
        self.logger = SinapiLogger("DatabaseManager", log_level)
    
    def log(self, level: str, message: str):
        """Wrapper para padronização de logs"""
        self.logger.log(level, f"[DB] {message}")
    
    @staticmethod
    def create_database(admin_connection_string: str, db_name: str, logger=None) -> bool:
        """
        Cria um banco de dados se não existir.
        Args:
            admin_connection_string (str): String de conexão com o banco de dados administrativo.
            db_name (str): Nome do banco de dados a ser criado.
            logger (optional): Objeto logger para registrar eventos.
        Returns:
            bool: True se o banco de dados foi criado com sucesso ou já existia, False se ocorreu um erro inesperado.
        """
        if logger is None:
            logger = SinapiLogger("DatabaseManager")
        logger.log('info', f"Verificando/criando banco de dados '{db_name}' usando conexão administrativa...")
        try:
            admin_engine = create_engine(admin_connection_string)
            with admin_engine.connect() as conn:
                conn.execution_options(isolation_level="AUTOCOMMIT").execute(
                    text(f"CREATE DATABASE {db_name}")
                )
            logger.log('info', f"Banco de dados '{db_name}' criado com sucesso.")
            admin_engine.dispose()
            return True
        except sqlalchemy.exc.ProgrammingError as e:
            if 'already exists' in str(e):
                logger.log('info', f"Banco de dados '{db_name}' já existe.")
                return True
            logger.log('error', f"Erro ao criar banco de dados '{db_name}': {e}")
            raise
        except Exception as e:
            logger.log('error', f"Erro inesperado ao criar banco de dados '{db_name}': {e}")
            return False

    def create_schemas(self, schemas: List[str]) -> None:
        """
        Cria esquemas se não existirem
        Args:
            schemas: Lista de esquemas
        """
        self.log('info', "Verificando esquemas...")
        for schema in schemas:
            try:
                with self.engine.connect() as conn:
                    conn.execute(text(f"CREATE SCHEMA IF NOT EXISTS {schema}"))
                    conn.commit()
                self.log('debug', f"Schema '{schema}' verificado")
            except Exception as e:
                self.log('error', f"Erro no schema '{schema}': {e}")
        self.log('info', f"Esquemas verificados: {', '.join(schemas)}")

    def _infer_sql_types(self, df: pd.DataFrame) -> List[str]:
        """Infere tipos SQL a partir dos dtypes do DataFrame"""
        type_mapping = []
        for col_name, dtype in df.dtypes.items():
            if 'int' in str(dtype):
                sql_type = "BIGINT"
            elif 'float' in str(dtype):
                sql_type = "NUMERIC(15,2)"
            else:
                sql_type = "TEXT"
            type_mapping.append(sql_type)
        return type_mapping

    def create_table(self, schema: str, table: str, df: pd.DataFrame = None, columns: List[str] = None, types: List[str] = None) -> None:
        """
        Cria uma tabela se não existir. Pode usar inferência de tipos a partir de um DataFrame.
        Args:
            schema: Nome do esquema
            table: Nome da tabela
            df: DataFrame para inferir colunas e tipos (opcional)
            columns: Lista de colunas (obrigatório se df não for fornecido)
            types: Lista de tipos SQL (obrigatório se df não for fornecido)
        """
        table_name = f"{schema}.{table}"
        self.log('info', f"Verificando tabela '{table_name}'...")
        
        if df is not None:
            # Modo inferência automática
            sql_types = self._infer_sql_types(df)
            col_defs = [f"{col} {typ}" for col, typ in zip(df.columns, sql_types)]
        elif columns and types:
            # Modo manual
            col_defs = [f"{col} {typ}" for col, typ in zip(columns, types)]
        if df is not None:
            sql_types = self._infer_sql_types(df)
            col_defs = [f'"{col}" {typ}' for col, typ in zip(df.columns, sql_types)] # Colunas com aspas
        elif columns and types:
            col_defs = [f'"{col}" {typ}' for col, typ in zip(columns, types)] # Colunas com aspas
        else:
            raise ValueError("Deve fornecer DataFrame ou colunas/tipos")
        
        ddl = f"CREATE TABLE IF NOT EXISTS {schema}.{table} ({', '.join(col_defs)})"
        # USA A NOVA FUNÇÃO PARA COMANDOS DDL
        self.execute_non_query(ddl)
        self.log('info', f"Tabela '{table}' criada/verificada")

    def insert_data(self, schema: str, table: str, df: pd.DataFrame, batch_size: int = 1000) -> None:
        """
        Insere dados em uma tabela com validação e progresso
        Args:
            schema: Nome do esquema
            table: Nome da tabela
            df: DataFrame com os dados
            batch_size: Tamanho do lote
        """
        table_name = f"{schema}.{table}"
        total_rows = len(df)
        start_time = time.time()
        
        self.log('info', f"Inserindo {total_rows:,} registros em {table_name}...")
        # ADIÇÃO: Converte todos os tipos de nulos (np.nan, pd.NA) para None, que o SQL entende como NULL.
        df = df.replace({pd.NA: None, np.nan: None})
        try:
            with self.engine.connect() as conn:
                with tqdm(total=total_rows, desc="Inserindo dados", unit="reg") as pbar:
                    for i in range(0, total_rows, batch_size):
                        batch_start = time.time()
                        batch_df = df.iloc[i:i + batch_size]
                        
                        try:
                            self._insert_batch(conn, table_name, batch_df)
                            conn.commit()
                        except Exception as e:
                            self.log('error', f"Erro no lote {i//batch_size + 1}: {e}")
                            continue
                        
                        batch_time = time.time() - batch_start
                        records = len(batch_df)
                        elapsed = time.time() - start_time
                        rate = records / batch_time if batch_time > 0 else 0
                        
                        pbar.update(records)
                        pbar.set_postfix({
                            "Tempo lote": f"{batch_time:.1f}s",
                            "Total": f"{elapsed:.1f}s",
                            "Reg/s": f"{rate:.0f}"
                        })
        except Exception as e:            
            self.log('error', f"Falha crítica durante a inserção de dados em {table_name}: {e}", exc_info=True)
            raise
        total_time = timedelta(seconds=int(time.time() - start_time))
        self.log('info', f"Inserção concluída em {total_time}")

    def _insert_batch(self, conn: sqlalchemy.engine.Connection, table_name: str, 
                     batch_df: pd.DataFrame) -> None:
        """
        Insere um lote de dados na tabela
        Args:
            conn: Conexão SQLAlchemy
            table_name: Nome completo da tabela (schema.tabela)
            batch_df: DataFrame com os dados do lote
        """
        data = batch_df.to_dict(orient='records')
        if not data:
            return
            
        columns = list(data[0].keys())
        placeholders = ', '.join([':' + col for col in columns])
        query = f"""
            INSERT INTO {table_name} ({', '.join(columns)})
            VALUES ({placeholders})
        """
        conn.execute(text(query), data)
    
    def _build_dynamic_query(self, base_query: str, filters: Dict[str, tuple]) -> list[str, Dict]:
        """
        Constroi query com cláusulas dinâmicas IN para múltiplas colunas.
        Args:
            base_query: Consulta base (sem WHERE)
            filters: Dicionário com {coluna: valores} para cláusulas IN
        Returns:
            Tuple: (query_completa, dicionário_de_parâmetros)
        """
        conditions = []
        params = {}
        for col, values in filters.items():
            if values:
                param_name = f"{col}_vals"
                conditions.append(f"{col} IN :{param_name}")
                params[param_name] = tuple(values)
        
        where_clause = " AND ".join(conditions) if conditions else "1=1"
        return f"{base_query} WHERE {where_clause}", params

    def _create_composite_key(self, df: pd.DataFrame, key_columns: List[str]) -> pd.Series:
        """Cria chave composta para comparação de registros"""
        # Preenche NA com string vazia para evitar NaN na chave
        return df[key_columns].fillna('').astype(str).apply('_'.join, axis=1)

    def validate_data(self, full_table_name: str, df: pd.DataFrame, backup_dir: Optional[Path] = None,policy: str = 'S') -> Optional[pd.DataFrame]:
        """
        Verifica dados existentes com política pré-definida
        Args:
            policy: Política para duplicatas (S=Substituir, A=Agregar, P=Pular)
        """
        schema_name, table_name = full_table_name.split('.')
        
        # Se tabela não existe, cria e retorna todos os dados
        if not self.table_exists(schema_name, table_name):
            self.create_table(schema_name, table_name, df=df)
            return df

        # Identificar colunas chave
        cod_columns = [col for col in df.columns if 'COD' in col]
        key_columns = cod_columns + ['ANO_REFERENCIA', 'MES_REFERENCIA'] if cod_columns else []
        
        # Sem colunas chave, retorna todos os dados
        if not key_columns:
            self.log('info', f"Não foram encontradas colunas chave para {table_name}")
            return df

        # Consultar registros existentes
        unique_values = {col: df[col].dropna().unique().tolist() for col in key_columns}
        base_query = f"SELECT {', '.join(key_columns)} FROM {full_table_name}"
        query, params = self._build_dynamic_query(base_query, unique_values)
        
        existing_df = self.execute_query(query, params)
        
        if existing_df.empty:
            self.log('info', f"Não foram encontradas duplicatas para {table_name}")
            return df

        # Verificar duplicatas
        df['temp_key'] = self._create_composite_key(df, key_columns)
        existing_df['temp_key'] = self._create_composite_key(existing_df, key_columns)
        
        df['is_duplicate'] = df['temp_key'].isin(existing_df['temp_key'])
        duplicates_count = df['is_duplicate'].sum()
        
        self.log('info', f"\nTabela: {table_name}")
        self.log('info', f"Total registros: {len(df)} | Duplicatas: {duplicates_count}")
        
        # Aplica política pré-definida
        choice = policy.upper()
        
        if choice == 'S':
            if backup_dir:
                self.backup_table(schema_name, table_name, Path(backup_dir))
            
            delete_query, delete_params = self._build_dynamic_query(
                f"DELETE FROM {full_table_name}", 
                unique_values
            )
            self.execute_non_query(delete_query, delete_params)
            self.log('info', f"Registros duplicados removidos: {duplicates_count}")
            df_to_insert = df.drop(columns=['temp_key', 'is_duplicate'])
        elif choice == 'A':
            df_to_insert = df[~df['is_duplicate']].drop(columns=['temp_key', 'is_duplicate'])
            self.log('info', f"Registros a inserir: {len(df_to_insert)}")
        else:
            self.log('info', "Inserção pulada por política.")
            return None
        
        return df_to_insert
    
    def execute_query(self, query: str, params: Dict = None, timeout: int = 30) -> pd.DataFrame:
        """
        Executa uma query SQL e retorna os resultados
        Args:
            query: Query SQL
            params: Parâmetros da query
            timeout: Timeout em segundos (opcional)
        Returns:
            DataFrame: Resultados da query
        """
        try:
            with self.engine.connect() as conn:
                result = conn.execute(text(query), params or {})
                if result.returns_rows:
                    return pd.DataFrame(result.fetchall(), columns=result.keys())
                else:
                    return pd.DataFrame()
        except Exception as e:
            self.log('error', f"Erro na query: {e}\nQuery: {query}")
            raise
    
    def execute_non_query(self, query: str, params: Dict = None, timeout: int = 30) -> None:
        """
        Executa um comando SQL que não retorna resultados (ex: CREATE, INSERT, DELETE, UPDATE).
        """
        try:
            with self.engine.connect() as conn:
                conn.execute(text(query), params or {})
                conn.commit() # Essencial para DDL e DML
        except Exception as e:
            self.log('error', f"Erro no comando non-query: {e}\nQuery: {query}")
            raise

    def backup_table(self, schema: str, table: str, backup_dir: Path) -> None:
        """
        Faz backup dos dados de uma tabela em CSV
        Args:
            schema: Nome do esquema
            table: Nome da tabela
            backup_dir: Diretório para salvar o backup
        """
        table_name = f"{schema}.{table}"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = backup_dir / f"{schema}_{table}_{timestamp}.csv"
        
        try:
            # Cria diretório se não existir
            backup_dir.mkdir(parents=True, exist_ok=True)
            
            # Recupera os dados
            query = f"SELECT * FROM {table_name}"
            df = self.execute_query(query)
            
            # Salva em CSV
            df.to_csv(backup_file, index=False)
            self.log('info', f"Backup salvo em {backup_file}")
            
        except Exception as e:
            self.log('error', f"Erro no backup de {table_name}: {e}")
            raise

    def get_connection_info(self) -> dict:
        """Retorna informações de conexão"""
        url = make_url(self.engine.url)
        return {
            'user': url.username,
            'host': url.host,
            'port': url.port,
            'dbname': url.database
        }
    
    def table_exists(self, schema: str, table: str,create: bool = False) -> bool: #inserir a criação de tabelas.
        """Verifica se tabela existe"""
        try:
            self.logger.log('info', f"[DB] Verificando tabela '{schema}.{table}'...")
            query = text(f"""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_schema = :schema 
                    AND table_name = :table
                )
            """)
            result = self.execute_query(query, {'schema': schema, 'table': table})
            return result.iloc[0, 0] if not result.empty else False
        except Exception as e:            
            self.log('error', f"Erro ao verificar existência da tabela {schema}.{table}: {e}")
            return False
    
    def optimize_inserts(self):
        """Configura otimizações para inserções em massa"""
        self.execute_query("SET session_replication_role = 'replica'")
    
    def restore_defaults(self):
        """Restaura configurações padrão do banco"""
        self.execute_query("SET session_replication_role = 'origin'")

class DatabaseConnection:
    """Gerenciador de contexto para conexão segura"""
    
    def __init__(self, secrets_path: str, log_level: str = 'info'):
        
        self.secrets_path = secrets_path
        self.log_level = log_level
        self.db_manager = None
    
    def __enter__(self) -> DatabaseManager:
        self.db_manager = create_db_manager(self.secrets_path, self.log_level)
        return self.db_manager
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        logger = SinapiLogger("DatabaseConnection")
        if self.db_manager:
            self.db_manager.engine.dispose()
        if exc_type:
            logger.log('error', f"Erro durante operação de banco: {exc_val}")
        return False
    
    @staticmethod
    def create_database_connection(params: dict) -> DatabaseManager:
        """
        Cria conexão com o banco de dados de forma otimizada
        
        Args:
            params: Dicionário de parâmetros contendo:
                - sql_secrets: Caminho para o arquivo de secrets
                - log_level: Nível de log
        
        Returns:
            Instância configurada de DatabaseManager
        """
        logger = SinapiLogger("DatabaseConnection")
        file_manager = FileManager()
        
        try:
            # Obtenção de credenciais
            creds = file_manager.read_sql_secrets(params['sql_secrets'])
            user, pwd, host, port, dbname, initial_db = creds
            
            # Teste de conexão
            test_conn_str = f"postgresql://{user}:{pwd}@{host}:{port}/{initial_db}"
            test_connection(test_conn_str)
            
            # Conexão principal
            main_conn_str = f"postgresql://{user}:{pwd}@{host}:{port}/{dbname}"
            db_manager = DatabaseManager(main_conn_str, params.get('log_level', 'info'))
            
            # Atualização de parâmetros
            conn_info = db_manager.get_connection_info()
            params.update({
                **conn_info,
                # Segurança penas certifique-se de que params não será usado para logar ou expor informações sensíveis em outro ponto do sistema.
                'password': '********',  
                'initial_db': initial_db
            })
            
            return db_manager
            
        except Exception as e:
            logger.log('error', f"Falha crítica na conexão com o banco: {e}")
            raise

class SinapiDownloader:
    """Classe para download de arquivos do SINAPI"""
    
    def __init__(self, cache_minutes: int = 10):
        self.logger = SinapiLogger("SinapiDownloader")
        self.cache_minutes = cache_minutes
        self.log_file = "sinap_webscraping_download_log.json"
    
    def _zip_pathfinder(self, folderName: str, ano: str, mes: str, formato: str = 'xlsx') -> str:
        folder_name = folderName
        zip_path = Path(folder_name) / f'SINAPI-{ano}-{mes}-formato-{formato}.zip'
        if zip_path.exists():
            self.logger.log('info', f'Arquivo já existe: {zip_path}')
            return str(zip_path)
        else:
            self.logger.log('info', f'Arquivo não existe: {zip_path}')
            return None

    def _zip_filefinder(self,folderName: str, ano: str, mes: str, formato: str = 'xlsx', dimiss: list = None, target: str = None) -> Union[str, tuple]:
        """
        Finds and selects ZIP files within a specified folder based on year, month, and format.
        Args:
            folderName (str): The name of the folder to search within.
            ano (str): The year to search for in the filename.
            mes (str): The month to search for in the filename.
            formato (str, optional): The file format to search for in the filename. Defaults to 'xlsx'.
            dimiss (list, optional): A list of filenames to exclude from the selection. Defaults to None.
            target (str, optional): The file name to search for in the files search result. Defaults to None
        Returns:
            tuple: A tuple containing two dictionaries:
                - zipFiles (dict): A dictionary where keys are filenames ending with '.zip' and values are their full paths.
                - selectFile (dict): A dictionary containing selected ZIP files based on the specified criteria,
                  excluding files present in the `dimiss` list (if provided). Keys are filenames and values are their full paths.
            None: Returns None if an error occurs during the file search.
            filepath: Return a filepath if file target is found.
        Raises:
            Exception: Logs any exceptions encountered during the file search process.
        """
        
        self.logger.log('info', f'Iniciando pesquisa do arquivo na pasta {folderName}')

        try:
            zipFiles = {}
            selectFile = {}
            path = Path(folderName)

            for arquivo in path.glob('*.zip'):
                zipFiles[arquivo.name] = str(arquivo)

            # Se target especificado, retorna caminho direto
            if target:
                return zipFiles.get(target)
            
                        
            # Filtra arquivos do SINAPI do mês/ano
            target_file = f'SINAPI-{ano}-{mes}-formato-{formato}.zip'
            selectFile = {k: v for k, v in zipFiles.items() if target_file in k}
            
            # Retorna consistente - caminho direto se apenas 1 arquivo, ou tupla se múltiplos
            if len(selectFile) == 1:
                return list(selectFile.values())[0]
            return zipFiles, selectFile
            
        except Exception as e:
            self.logger.log('error', f'Erro ao buscar arquivos: {str(e)}')
            return None
                        
    def download_file(self, ano: str, mes: str, formato: str = 'xlsx',sleeptime: int = 2, count: int = 4) -> Optional[str]:
        """
        Baixa arquivo do SINAPI se necessário
        Args:
            ano (str): Ano de referência (YYYY)
            mes (str): Mês de referência (MM)
            formato (str): Formato do arquivo ('xlsx' ou 'pdf')
            sleeptime (int): Tempo de espera entre tentativas de download
        Returns:
            Optional[str]: Caminho do arquivo baixado ou None se falhou
        """
        if not self._validar_parametros(ano, mes, formato):
            return None
            
        if not self._pode_baixar(ano, mes):
            return None
            
        url = f'https://www.caixa.gov.br/Downloads/sinapi-relatorios-mensais/SINAPI-{ano}-{mes}-formato-{formato}.zip'
        folder_name = f'./{ano}_{mes}'
        
        #print('iniciando pesquisa...')
        zip_path = self._zip_pathfinder(folder_name,ano,mes,formato)
        
        if zip_path:
            return str(zip_path)
        
        try:
            os.makedirs(folder_name, exist_ok=True)
            try:
                try:
                    download = self._download_with_retry(url, zip_path,retry_delays = [10, 30, 60], timeout=120)
                    self.logger.log('info', f'Download concluído: {zip_path}')
                    if download is True:
                        self._registrar_download(ano, mes)
                        return str(zip_path)
                    else:
                        raise
                except Exception as e:
                    self.logger.log('warning', f'Primeira tentativa de download falhou: {str(e)}')
                    self.logger.log('info', f'\nTentando {count} downloads com proxies...')
                    self._download_with_proxies(url, zip_path, str(ano), str(mes), int(sleeptime),count)
                    self._registrar_download(ano, mes)
                    return str(zip_path)
            
            except Exception as e:
                self.logger.log('error', f'Erro no download: {str(e)}')
                return None
            
            
        except Exception as e:
            self.logger.log('error', f'Erro no download: {str(e)}')
            return None

    def _download_with_proxies(self, url: str, zip_path: Path, ano: str, mes: str,sleeptime: int,count: int = 0) -> None:
        """
        Baixa um arquivo usando uma lista de proxies públicos.
        
        Args:
            url (str): URL do arquivo a ser baixado.
            zip_path (Path): Caminho local para salvar o arquivo.
            ano (str): Ano de referência (YYYY)
            mes (str): Mês de referência (MM)
        
        Raises:
            Exception: Se o download falhar após tentar vários proxies.
        """
        proxies_url = "https://cdn.jsdelivr.net/gh/proxifly/free-proxy-list@main/proxies/all/data.json"
        try:
            proxies_resp = requests.get(proxies_url, timeout=30)
            proxies_resp.raise_for_status()
            proxies_list = proxies_resp.json()
            random.shuffle(proxies_list)
        except Exception as e:
            raise Exception(f'Erro ao obter lista de proxies: {str(e)}') from e

        success = False
        self.logger.log('info', f'Tentando baixar com {len(proxies_list)} proxies...')
        if count == 0:
            count = len(proxies_list)
        else:
            for i,attempt in enumerate(range(count)):
                self.logger.log('info', f'\n=============================\n    >>>>>>> Tentativa nª{i+1} / {len(proxies_list)} <<<<<<<\n')
                proxy_info = random.choice(proxies_list)
                proxy = proxy_info.get("proxy")
                if not proxy:
                    self.logger.log('warning', 'Proxy não encontrado na lista, pulando.')
                    continue

                proxies = {
                    "http": f"http://{proxy}",
                    "https": f"http://{proxy}",
                }

                try:
                    self.logger.log('info', f'Tentando download com proxy: {proxy}')
                    session = requests.Session()
                    adapter = requests.adapters.HTTPAdapter(max_retries=1)
                    session.mount('https://', adapter)
                    response = session.get(url, timeout=120, allow_redirects=True, proxies=proxies)
                    response.raise_for_status()
                    with open(zip_path, 'wb') as f:
                        f.write(response.content)
                    self.logger.log('info', f'Download concluído com proxy: {proxy}')
                    success = True
                    break  # Encerra o loop assim que um proxy funciona
                except Exception as e:
                    self.logger.log('warning', f'Falha com proxy {proxy}: {str(e)}\n')
                    time.sleep(sleeptime)  # Adiciona um pequeno delay antes de tentar o próximo proxy

            if not success:
                raise Exception('Não foi possível baixar o arquivo com nenhum proxy')

            self._registrar_download(ano, mes)
            return str(zip_path)

    def unzip_file(self, zip_path: Union[str, Path]) -> Optional[str]:
        """
        Extrai um arquivo ZIP do SINAPI
        Args:
            zip_path: Caminho do arquivo ZIP
        Returns:
            Optional[str]: Caminho da pasta extraída ou None se falhou
        """
        zip_path = Path(zip_path)
        if not zip_path.exists():
            self.logger.log('error', f'Arquivo não existe: {zip_path}')
            return None
            
        extraction_path = zip_path.parent / zip_path.stem
        
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(extraction_path)
            self.logger.log('info', f'Arquivos extraídos em: {extraction_path}')
            return str(extraction_path)
        except Exception as e:
            self.logger.log('error', f'Erro ao extrair: {str(e)}')
            return None

    def _validar_parametros(self, ano: str, mes: str, formato: str) -> bool:
        """Valida os parâmetros de entrada"""
        try:
            if len(ano) != 4 or len(mes) != 2:
                raise ValueError("Ano deve ter 4 dígitos e mês deve ter 2 dígitos")
            if int(mes) < 1 or int(mes) > 12:
                raise ValueError("Mês deve estar entre 01 e 12")
            if formato not in ['xlsx', 'pdf']:
                raise ValueError("Formato deve ser 'xlsx' ou 'pdf'")
            return True
        except Exception as e:
            self.logger.log('error', f'Parâmetros inválidos: {str(e)}')
            return False

    def _pode_baixar(self, ano: str, mes: str) -> bool:
        """Verifica se já passou o tempo mínimo desde o último download"""
        chave = f"{ano}_{mes}"
        agora = datetime.now()
        
        if not os.path.exists(self.log_file):
            return True
            
        try:
            with open(self.log_file, "r") as f:
                log = json.load(f)
            ultimo = log.get(chave)
            if ultimo:
                ultimo_dt = datetime.fromisoformat(ultimo)
                if agora - ultimo_dt < timedelta(minutes=self.cache_minutes):
                    tempo_restante = timedelta(minutes=self.cache_minutes) - (agora - ultimo_dt)
                    self.logger.log('warning', 
                        f"Download recente detectado. Aguarde {tempo_restante} antes de tentar novamente.")
                    return False
        except Exception as e:
            self.logger.log('error', f'Erro ao ler log: {str(e)}')
            
        return True

    def _registrar_download(self, ano: str, mes: str) -> None:
        """Registra a data/hora do download no log"""
        chave = f"{ano}_{mes}"
        agora = datetime.now().isoformat()
        log = {}
        
        if os.path.exists(self.log_file):
            try:
                with open(self.log_file, "r") as f:
                    log = json.load(f)
            except Exception:
                pass
                
        log[chave] = agora
        
        with open(self.log_file, "w") as f:
            json.dump(log, f)

    def _download_with_retry(self, url: str, zip_path: Path, retry_delays: list = [10, 30, 60], timeout: int = 120) -> None:
        """Faz o download com retry em caso de falha com delays configuráveis
            Args:
                url (str): URL do arquivo a ser baixado.
                zip_path (Path): Caminho local para salvar o arquivo baixado.
                retry_delays (list, optional): Lista de tempos de espera em segundos entre cada tentativa. Defaults to [10, 30, 60].
                timeout (int, optional): Tempo máximo em segundos para aguardar uma resposta do servidor. Defaults to 120.
            Returns:
                bool: True se o download for bem-sucedido.
            Raises:
                requests.exceptions.HTTPError: Se ocorrer um erro HTTP durante o download.
                Exception: Se ocorrer qualquer outro erro durante o download.
            """

        session = requests.Session()
        adapter = requests.adapters.HTTPAdapter(max_retries=len(retry_delays))
        session.mount('https://', adapter)

        for attempt, delay in enumerate(retry_delays):
            try:
                self.logger.log('info', f'Tentativa de download {attempt + 1} de {len(retry_delays)}, aguardando {delay} segundos...')
                time.sleep(delay)
                response = session.get(url, timeout=timeout, allow_redirects=True)
                response.raise_for_status()
                
                with open(zip_path, 'wb') as f:
                    f.write(response.content)
                self.logger.log('info', f'Download concluído: {zip_path}')
                return True
                
            except requests.exceptions.HTTPError as e:
                if e.response.status_code == 404:
                    self.logger.log('error', 'Arquivo não encontrado no servidor')
                else:
                    self.logger.log('error', f'Erro HTTP na tentativa {attempt + 1}: {str(e)}\n')
                
                
            except Exception as e:
                self.logger.log('error', f'Erro no download na tentativa {attempt + 1}: {str(e)}\n')
                if attempt == len(retry_delays) - 1:
                    raise  # Relevanta a exceção na última tentativa

        return False  # Retorna False se todas as tentativas falharem

class SinapiProcessor:
    """Classe para processamento específico das planilhas SINAPI"""
    
    def __init__(self):
        self.logger = SinapiLogger("SinapiProcessor")
        self.file_manager = FileManager()
    
    def process_excel(self, file_path: Union[str, Path], sheet_name: str, header_id: int, split_id: int = 0) -> pd.DataFrame:
        """
        Processa uma planilha SINAPI, normalizando colunas e realizando transformações necessárias
        Args:
            file_path: Caminho do arquivo Excel
            sheet_name: Nome da planilha
            header_id: Índice da linha de cabeçalho
            split_id: Índice para split de colunas (melt)
        Returns:
            DataFrame: Dados processados
        """
        #iniciando
        self.logger.log('info', f'Processando planilha {sheet_name} do arquivo {file_path}')
        try:            
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_id)
            #self.logger.log('info', f'{df.head()}')
            
            # Normaliza nomes das colunas
            df.columns = [self.file_manager.normalize_text(col) for col in df.columns]
            #self.logger.log('info', f'Colunas: {list(df.columns)}')
            
            # Se necessário fazer melt (unpivot)
            if split_id > 0:
                self.logger.log('debug', f'Realizando melt com split_id={split_id}')
                df = pd.melt(
                    df,
                    id_vars=df.columns[:split_id],
                    value_vars=df.columns[split_id:],
                    var_name='ESTADO',
                    value_name='COEFICIENTE'
                    )
                #self.logger.log('info', f'{df.head()}')
            return df
            
        except Exception as e:
            self.logger.log('error', f'Erro ao processar planilha: {str(e)}')
            raise
            
    def identify_sheet_type(self, sheet_name: str, table_names: List[str] = None) -> Dict[str, int]:
        """
        Identifica o tipo de planilha SINAPI e retorna suas configurações
        Args:
            sheet_name: Nome da planilha
            table_names: Lista de nomes de tabelas conhecidas
        Returns:
            Dict: Configurações da planilha (split_id, header_id)
        """
        sheet_name = self.file_manager.normalize_text(sheet_name)
        
        # Configurações padrão por tipo de planilha
        configs = {
            'ISD': {'split_id': 5, 'header_id': 9},
            'CSD': {'split_id': 4, 'header_id': 9},
            'ANALITICO': {'split_id': 0, 'header_id': 9},
            'COEFICIENTES': {'split_id': 5, 'header_id': 5},
            'MANUTENCOES': {'split_id': 0, 'header_id': 5},
            'MAO_DE_OBRA': {'split_id': 4, 'header_id': 5}
        }
        
        # Verifica correspondências diretas
        for type_name, config in configs.items():
            if type_name in sheet_name:
                self.logger.log('info', f'Planilha {sheet_name} identificada como {type_name}')
                return config
        
        # Verifica correspondências com table_names se fornecido
        if table_names:
            for i, table in enumerate(table_names):
                table = self.file_manager.normalize_text(table)
                if table in sheet_name:
                    if i == 0:  # Insumos Coeficiente
                        return {'split_id': 5, 'header_id': 5}
                    elif i == 1:  # Códigos Manutenções
                        return {'split_id': 0, 'header_id': 5}
                    elif i == 2:  # Mão de Obra
                        return {'split_id': 4, 'header_id': 5}
        
        self.logger.log('warning', f'Tipo de planilha não identificado: {sheet_name}')
        return None
    
    def validate_data(self, df: pd.DataFrame, expected_columns: List[str] = None) -> bool:
        """
        Valida os dados de uma planilha SINAPI
        Args:
            df: DataFrame a ser validado
            expected_columns: Lista de colunas esperadas
        Returns:
            bool: True se válido, False caso contrário
        """
        try:
            # Verifica se há dados
            if df.empty:
                self.logger.log('error', 'DataFrame está vazio')
                return False
            
            # Verifica colunas esperadas
            if expected_columns:
                missing = set(expected_columns) - set(df.columns)
                if missing:
                    self.logger.log('error', f'Colunas ausentes: {missing}')
                    return False
            
            # Verifica valores nulos em colunas críticas
            critical_cols = [col for col in df.columns if 'COD' in col or 'ID' in col]
            for col in critical_cols:
                null_count = df[col].isnull().sum()
                if null_count > 0:
                    self.logger.log('warning', f'Coluna {col} tem {null_count} valores nulos')
            
            return True
            
        except Exception as e:
            self.logger.log('error', f'Erro na validação: {str(e)}')
            return False
    
    def clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Limpa e padroniza os dados de uma planilha SINAPI
        Args:
            df: DataFrame a ser limpo
        Returns:
            DataFrame: Dados limpos e padronizados
        """
        try:
            df = df.copy()
            
            # Remove linhas totalmente vazias
            df = df.dropna(how='all')
            
            # Limpa strings
            str_columns = df.select_dtypes(include=['object']).columns
            for col in str_columns:
                df[col] = df[col].apply(lambda x: self.file_manager.normalize_text(str(x)) if pd.notnull(x) else x)
            
            # Converte colunas numéricas
            num_columns = df.select_dtypes(include=['float64', 'int64']).columns
            for col in num_columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # Remove caracteres especiais de colunas específicas
            cod_columns = [col for col in df.columns if 'CODIGO' in col]
            for col in cod_columns:
                self.logger.log('info',f'Coluna Código Encontrada:  {col}')
                df[col] = df[col].astype(str).str.replace(r'[^0-9]', '', regex=True)

            desc_columns = [col for col in df.columns if 'DESCRICAO' in col]
            for col in desc_columns:
                self.logger.log('info',f'Coluna Descrição Encontrada: {col}')
                df[col] = df[col].astype(str).str.replace('  ',' ', regex=True).apply(
                lambda x: self.file_manager.normalize_text(x) if x else x
            ).str.replace('_',' ', regex=True)
                #df[col] = df[col].astype(str).str.replace('_',' ', regex=True)

            return df
            
        except Exception as e:
            self.logger.log('error', f'Erro na limpeza dos dados: {str(e)}')
            raise

class SinapiPipeline:
    def __init__(self, config_path="CONFIG.json"):
        self.config = self.load_config(config_path)
        self.logger = SinapiLogger("SinapiPipeline", self.config.get('log_level', 'info'))
        self.downloader = SinapiDownloader(cache_minutes=90)
        self.file_manager = FileManager()
        self.processor = SinapiProcessor()
        self.excel_processor = ExcelProcessor()
        self.db_manager = None

    def load_config(self, config_path):
        """Carrega o arquivo de configuração"""
        try:
            with open(config_path, 'r') as f:
                return json.load(f)
        except Exception as e:
            self.logger.log('critical', f'Erro ao carregar configuração: {e}')
            raise

    def get_parameters(self):
        """Obtém parâmetros de execução do config ou do usuário"""
        params = {
            'ano': self.config.get('default_year'),
            'mes': self.config.get('default_month'),
            'formato': self.config.get('default_format', 'xlsx'),
            'workbook': self.config.get('default_workbook', 'REFERENCIA'),
        }
        
        # Preenche valores faltantes via input do usuário
        if not params['ano']:
            params['ano'] = input('Digite o ano (YYYY): ').strip()
        if not params['mes']:
            params['mes'] = input('Digite o mês (MM): ').strip()
        
        # Validação básica
        if not params['ano'].isdigit() or len(params['ano']) != 4:
            raise ValueError("Ano inválido!")
        if not params['mes'].isdigit() or len(params['mes']) != 2:
            raise ValueError("Mês inválido!")
        
        return params

    def setup_database(self):
        """Configura a conexão com o banco de dados"""
        secrets_path = self.config['secrets_path']
        
        if not os.path.exists(secrets_path):
            raise FileNotFoundError(f'Arquivo de secrets não encontrado: {secrets_path}')
        
        self.db_manager = create_db_manager(
            secrets_path=secrets_path,
            log_level=self.config.get('log_level', 'info'),
            output='target'
        )
        
        # Garante existência dos schemas
        self.db_manager.create_schemas(['public', 'sinapi'])

    def download_and_extract_files(self, params):
        """Gerencia download e extração de arquivos"""
        ano = params['ano']
        mes = params['mes']
        formato = params['formato']
        diretorio_referencia = f"./{ano}_{mes}"
        
        # Cria diretório se necessário
        if not os.path.exists(diretorio_referencia):
            os.makedirs(diretorio_referencia)
        
        # Verifica se o arquivo já existe
        filefinder_result = self.downloader._zip_filefinder(diretorio_referencia, ano, mes, formato)
        
        # Trata diferentes tipos de retorno
        if isinstance(filefinder_result, tuple) and len(filefinder_result) == 2:
            # Caso retorne (zipFiles, selectFile)
            zip_files, select_file = filefinder_result
            if select_file:
                zip_path = list(select_file.values())[0]
            else:
                zip_path = None
        elif isinstance(filefinder_result, str):
            # Caso retorne diretamente o caminho
            zip_path = filefinder_result
        else:
            zip_path = None
        
        # Faz download se necessário
        if not zip_path:
            zip_path = self.downloader.download_file(ano, mes, formato, sleeptime=1, count=3)
            if not zip_path:
                raise RuntimeError("Falha no download do arquivo")
        
        # Extrai arquivos
        extraction_path = self.downloader.unzip_file(zip_path)
        return extraction_path

    def process_spreadsheets(self, extraction_path, planilha_name):
        """Processa as planilhas usando padrão Strategy"""
        # Normaliza nomes de arquivos
        self.file_manager.normalize_files(extraction_path, extension='xlsx')
        
        # Identifica planilhas disponíveis
        workbook = self.excel_processor.scan_directory(
            diretorio=extraction_path,
            formato='xlsx',
            data=True,
            sheet={planilha_name: extraction_path}
        )
        
        df_list = {}
        for sheet_name in workbook.get(planilha_name, []):
            sheet_config = self.config['sheet_processors'].get(
                sheet_name[0],
                self.processor.identify_sheet_type(sheet_name[0])
            )
            
            if not sheet_config:
                self.logger.log('warning', f'Configuração não encontrada para: {sheet_name[0]}')
                continue
                
            # Processa a planilha
            df = self.processor.process_excel(
                f'{extraction_path}/{planilha_name}',
                sheet_name[0],
                sheet_config['header_id'],
                sheet_config['split_id']
            )
            
            # Limpa e normaliza dados
            clean_df = self.processor.clean_data(df)
            table_name = self.file_manager.normalize_text(
                f'{planilha_name.split(".")[0]}_{sheet_name[0]}'
            )
            df_list[table_name] = clean_df
        
        return df_list

    def handle_database_operations(self, df_list):
        """Gerencia operações de banco de dados"""
        duplicate_policy = self.config.get('duplicate_policy', 'substituir').upper()
        backup_dir = self.config.get('backup_dir', './backups')
        
        for table_name, df in df_list.items():
            full_table_name = f"sinapi.{table_name}"
            
            try:
                # Valida e prepara dados para inserção
                df_to_insert = self.db_manager.validate_data(
                    full_table_name=full_table_name,
                    df=df,
                    backup_dir=backup_dir,
                    policy=duplicate_policy  # Novo parâmetro
                )
                
                if df_to_insert is not None:
                    self.db_manager.insert_data('sinapi', table_name, df_to_insert)
            except Exception as e:
                self.logger.log('error', f'Erro ao inserir dados: {e}')

    def export_results(self, df_list, params):
        """Exporta resultados para CSV"""
        ano = params['ano']
        mes = params['mes']
        diretorio_referencia = f"./{ano}_{mes}"
        
        for table_name, df in df_list.items():
            csv_filename = f"{table_name}_{ano}_{mes}.csv"
            csv_path = os.path.join(diretorio_referencia, csv_filename)
            df.to_csv(csv_path, index=False, encoding='utf-8-sig')
            self.logger.log('info', f"DataFrame exportado para: {csv_path}")

#Funções Auxiliares

logger = SinapiLogger("test_connection")

def create_db_manager(secrets_path: str, log_level: str = 'info',output: str = None) -> DatabaseManager:
    """
    Cria e retorna um DatabaseManager configurado a partir de arquivo de secrets
    
    Args:
        secrets_path: Caminho completo para o arquivo .secrets
        log_level: Nível de log desejado
        output:  Tipo de saída desejada:
            test: dbtest
            target: dbtarget
        
    Returns:
        Instância configurada de DatabaseManager
        
    Raises:
        FileNotFoundError: Quando arquivo não existe
        ConnectionError: Quando falha teste de conexão
        ValueError: Quando credenciais incompletas
    """
    logger = SinapiLogger("create_db_manager", log_level)

    # Validação inicial do arquivo
    logger.log('info', f"Iniciando criação do DatabaseManager com arquivo de secrets: {secrets_path}")
    if not os.path.isfile(secrets_path):
        logger.log('error', f"Arquivo de secrets não encontrado: {secrets_path}")
        raise FileNotFoundError(f"Arquivo de secrets não encontrado: {secrets_path}")
    
    # Leitura das credenciais
    try:
        logger.log('debug', f"Lendo credenciais do arquivo: {secrets_path}")
        with open(secrets_path, 'r') as f:
            content = f.read()
        credentials = parse_secrets(content)
        logger.log('debug', "Credenciais lidas com sucesso.")
    except Exception as e:
        logger.log('critical', f"Falha na leitura do arquivo secrets: {e}")
        raise RuntimeError(f"Falha na leitura do arquivo secrets: {e}") from e

    # Validação mínima das credenciais
    required_keys = {'DB_USER', 'DB_PASSWORD', 'DB_HOST', 'DB_PORT', 'DB_NAME'}
    if not required_keys.issubset(credentials):
        missing = required_keys - set(credentials.keys())
        logger.log('error', f"Credenciais incompletas. Faltando: {', '.join(missing)}")
        raise ValueError(f"Credenciais incompletas. Faltando: {', '.join(missing)}")

    # Teste de conexão com banco inicial
    user=credentials['DB_USER']
    pwd=credentials['DB_PASSWORD']
    host=credentials['DB_HOST']
    port=credentials['DB_PORT']
    initial_db=credentials.get('DB_INITIAL_DB', 'postgres') # Definir a base a ser utilizada
    test_conn_str = f"postgresql://{user}:{pwd}@{host}:{port}/{initial_db}"
    
    try:
        logger.log('info', f"Testando conexão com banco inicial: {host}:{port}/{initial_db}")
        test_connection(test_conn_str)
        logger.log('info', f"Teste de conexão com banco inicial ({initial_db}) bem-sucedido.\n")
    except Exception as e:
        logger.log('critical', f"Falha no teste de conexão com banco inicial: {e}")
        raise ConnectionError(f"Falha na conexão com banco inicial: {e}") from e
    
    #verifica qual o tipo de saída desejada:
    if output == 'test':
        target_db = credentials['DB_INITIAL_DB']
    elif output == 'target':
        target_db = credentials['DB_NAME']
    elif output == None:
        target_db = 'postgres'
    else:
        logger.log('error', f"Tipo de saída inválido: {output}")
        raise ValueError(f"Tipo de saída inválido: {output}")
    
    try:
        conn_str = format_connection_string(
            user=credentials['DB_USER'],
            password=credentials['DB_PASSWORD'],
            host=credentials['DB_HOST'],
            port=credentials['DB_PORT'],
            database=target_db
        )
        test_connection(conn_str)
    except Exception as e:
        logger.log('warning', f"Falha ao conectar ao banco de dados '{target_db}': {e}")
        
        # Tenta criar o banco de dados se a conexão falhar
        admin_conn_str = format_connection_string(
            user=credentials['DB_USER'],
            password=credentials['DB_PASSWORD'],
            host=credentials['DB_HOST'],
            port=credentials['DB_PORT'],
            database=initial_db  # Usa o banco de dados inicial para criar o novo
        )
        
        try:
            logger.log('info', f"Tentando criar o banco de dados '{target_db}'...")
            DatabaseManager.create_database(admin_conn_str, target_db, logger)
            logger.log('info', f"Banco de dados '{target_db}' criado com sucesso.")
            
            # Recria a string de conexão após a criação do banco
            conn_str = format_connection_string(
                user=credentials['DB_USER'],
                password=credentials['DB_PASSWORD'],
                host=credentials['DB_HOST'],
                port=credentials['DB_PORT'],
                database=target_db
            )
            test_connection(conn_str)
            
        except Exception as create_err:
            logger.log('critical', f"Falha ao criar o banco de dados '{target_db}': {create_err}")
            raise

    # Cria e retorna o DatabaseManager
    logger.log('info', f"Criando DatabaseManager com conexão: {host}:{port}/{target_db}")
    db_manager = DatabaseManager(
        connection_string=conn_str,
        log_level=log_level
    )
    
    logger.log('info', f"DatabaseManager criado com sucesso. ({conn_str})\n")
    return db_manager

def format_connection_string(user: str, password: str, host: str, port: str, database: str) -> str:
    """Formata a string de conexão de forma segura (usa SQLAlchemy URL para evitar injeção)"""
    return URL.create(
        drivername="postgresql",
        username=user,
        password=password,
        host=host,
        port=port,
        database=database
    ).render_as_string(hide_password=False)

def test_connection(conn_str: str, timeout: int = 5, retries: int = 3):
    """Testa conexão com retentativas automáticas"""
    logger.log('info', f"Testando conexão com timeout={timeout}s e {retries} retentativas.")
    for attempt in range(retries):
        try:
            logger.log('debug', f"Tentativa de conexão {attempt + 1}/{retries}...")
            engine = create_engine(conn_str, connect_args={'connect_timeout': timeout})
            with engine.connect() as conn:
                conn.execute(text("SELECT 1"))
            logger.log('info', "Conexão bem-sucedida!")
            return
        except Exception as e:
            logger.log('warning', f"Falha na tentativa {attempt + 1}: {e}")
            if attempt < retries - 1:
                sleep(2 ** attempt)  # Backoff exponencial
                logger.log('debug', f"Aguardando {2 ** attempt} segundos antes da próxima tentativa...")
            else:
                logger.log('error', f"Falha na conexão após {retries} tentativas.")
                raise ConnectionError(f"Falha na conexão após {retries} tentativas: {e}")
            
def parse_secrets(content: str) -> dict:
    """Extrai credenciais do conteúdo do arquivo secrets"""
    credentials = {}
    for line in content.splitlines():
        match = re.match(r"\s*([A-Z_]+)\s*=\s*'([^']*)'", line)
        if match:
            key, value = match.groups()
            credentials[key] = value
    return credentials

def make_conn_str(info: dict, dbname: str = None):
    """Gera uma connection string a partir do dicionário de informações e nome do banco."""
    return f"postgresql://{info['user']}:{info['password']}@{info['host']}:{info['port']}/{dbname or info['dbname']}"
