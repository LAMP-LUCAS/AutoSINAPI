'''import pandas as pd
import os
import sinapi_utils

planilhas = ['familias_e_coeficientes', 'Manutenções', 'mao_de_obra', 'Referência']
formatos = ['xlsx', 'pdf']
ano = '2025'
mes = '03'


# 1. Carregar o arquivo Excel
file_path = '{ano}_{mes}/SINAPI-{ano}{mes}-formato-{formato}/SINAPI_{planilha}_{ano}_{mes}.xlsx'
df = pd.read_excel(file_path)

# 2. Identificar colunas fixas e colunas de estados
fixed_columns = ['mes/ref', 'cod familia', 'cod insumo', 'descr', 'unid', 'categoria']
estados_columns = [col for col in df.columns if col not in fixed_columns]

# 3. Realizar o 'unpivot' (melt) para normalizar os dados
df_normalized = pd.melt(
    df,
    id_vars=fixed_columns,
    value_vars=estados_columns,
    var_name='estado',
    value_name='coeficiente'
)

# 4. Renomear colunas para padrão SQL
df_normalized = df_normalized.rename(columns={
    'mes/ref': 'mes_ref',
    'cod familia': 'cod_familia',
    'cod insumo': 'cod_insumo',
    'descr': 'descricao',
    'unid': 'unidade'
})

# 5. Tratamento de dados (opcional, mas recomendado)
# - Converter tipo de data (se necessário)
# df_normalized['mes_ref'] = pd.to_datetime(df_normalized['mes_ref'], format='%m/%Y')

# - Remover linhas com coeficientes nulos
df_normalized = df_normalized.dropna(subset=['coeficiente'])

# - Converter estado para maiúsculas
df_normalized['estado'] = df_normalized['estado'].str.upper()

# 6. Salvar para CSV (para posterior importação via COPY)
output_csv = 'dados_normalizados.csv'
df_normalized.to_csv(output_csv, index=False, encoding='utf-8')

print(f"Normalização concluída! Dados salvos em: {output_csv}")
print(f"Total de linhas originais: {len(df)}")
print(f"Total de linhas normalizadas: {len(df_normalized)}")
print(f"Exemplo de dados normalizados:")
print(df_normalized.head())

# ---------------------------------------------------------
# 7. (OPCIONAL) Inserção direta no PostgreSQL
# Requer instalação: pip install sqlalchemy psycopg2

from sqlalchemy import create_engine

# Configurações do banco
DB_USER = 'seu_usuario'
DB_PASSWORD = 'sua_senha'
DB_HOST = 'localhost'
DB_PORT = '5432'
DB_NAME = 'seu_banco'
TABLE_NAME = 'sinapi_insumos'

# Criar conexão
engine = create_engine(f'postgresql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}')

# Inserir dados diretamente
# df_normalized.to_sql(
#     name=TABLE_NAME,
#     con=engine,
#     if_exists='append',  # ou 'replace' para recriar a tabela
#     index=False,
#     method='multi'  # inserção em lote
# )

# print(f"\nDados inseridos diretamente na tabela {TABLE_NAME}!")


'''

import os
import pandas as pd
import sinapi_utils
from sinapi_utils import normalize_text
import sinap_webscraping
import rastreador_xlsx
from datetime import datetime
from openpyxl import load_workbook

def main():
    try:
        # perguntar o ano e mes de referencia da base sinapi para insersão no banco de dados

        data_ano_referencia = input(f'Digite o ano (YYYY): ')
        data_mes_referencia = input(f'Digite o mes (MM): ')
        formato_arquivos = input(f'Digite o formato dos arquivos (1 - xlsx, 2 - pdf): ')

        # Tratando mes e ano
        if data_ano_referencia:
            try:
                if len(data_ano_referencia) != 4:
                    print(f'Ano inválido')
                    exit()
                if int(data_ano_referencia) < 2025:
                    tipoBase = 2025
                    print(f'Base utilizada à partir de 2025')
                ano = data_ano_referencia

            except ValueError:
                print(f'Valor {data_ano_referencia} inválido para o ano')

        if data_mes_referencia:
            try:
                if len(data_mes_referencia) < 2:
                    print(f'Mês inválido')
                    exit()
                if int(data_mes_referencia) > 12:
                    print(f'Mês inválido')
                mes = data_mes_referencia
            except ValueError:
                print(f'Valor {data_mes_referencia} inválido para o mes')

        if formato_arquivos:
            try:
                if formato_arquivos in ['1', '2']: 
                    if formato_arquivos == '1':
                        formato = 'xlsx'
                    elif formato_arquivos == '2':
                        formato = 'pdf'
                elif formato_arquivos in ['xlsx', 'pdf']:
                    if formato_arquivos == 'xlsx':
                        formato = 'xlsx'
                    elif formato_arquivos == 'pdf':
                        formato = 'pdf'
            except ValueError:
                    print(f'Valor {formato_arquivos} inválido para o formato do arquivo')

        # verificando se arquivos existem na pasta e listando-os ou perguntando se o usuário quer fazer o download da base.
        diretorio_atual = os.getcwd()
        diretorio = f'{diretorio_atual}/{ano}_{mes}/SINAPI-{ano}-{mes}-formato-{formato}'
        diretorio = diretorio.replace('/', '\\')
        resultado = rastreador_xlsx.scan_excel_directory(diretorio)

        if resultado == None:
            print('Não foram encontrados arquivos no diretório: ', diretorio)
            exit()

        planilhas = list(resultado.keys())

        #print(f'planilhas: {planilhas}')

        #input do usuário para definir qual planilha utilizar:

        print(f'\nEscolha a planilha que deseja utilizar:')
        menu_planilhas = [print(f'    {i} - {planilhas[i]}') for i in range(len(planilhas))]

        planilha_index = input(f'Escolha ao índice da planilha que deseja utilizar: ')
        planilha = planilhas[int(planilha_index)]
        print(f'\nPlanilha escolhida: {planilha_index} - {planilha}\n')

        file_path = f'{diretorio}/{planilha}'
        file_path = file_path.replace('\\', '/')

        planilha_name = str(planilha).split('SINAPI_')[-1].split(f'_{ano}_{mes}.')[0]
        planilha_name = planilha_name.replace('_', ' ').replace('-', ' ').upper()

        tableNames = ('COEFICIENTES', 'MANUTENCOES', 'SEM_DESONERACAO', 'ANALITICO', 'PRECOS')

        #print(f'    Planilha-Name: {planilha_name}')

        # lista todas as planilhas do arquivo excel e verifica se o planilha_name tem relação com algum nome
        # Lista todas as planilhas do arquivo Excel
        workbook = load_workbook(file_path, read_only=True)
        all_sheets = workbook.sheetnames
        print(f"\n    Planilhas disponíveis no arquivo: {all_sheets}\n    TableNames: {tableNames}\n")

        # Verifica se planilha_name tem relação com alguma das planilhas do arquivo
        matched_sheet = [] if 'REFERENCIA' in planilha_name else None
        tableReferenciaName = [] if 'REFERENCIA' in planilha_name else None

        for sheet in all_sheets:
            normalized_sheet = normalize_text(sheet.upper())
            print(f'    Planilha: {normalized_sheet}')
            if isinstance(matched_sheet, list):
                tableReferenciaName = ['ISD','CSD','ANALITICO']

            # Se for lista, verificar se todos os itens de tableNames estão na sheet
            if all(item in normalized_sheet for item in tableReferenciaName):
                matched_sheet.append(sheet)
                print(f'        Todas as tabelas {tableReferenciaName} encontradas em: {sheet}')
            else:
                # Caso contrário, procurar por cada tabela individualmente
                for table in tableNames:
                    if table in normalized_sheet:
                        matched_sheet = sheet
                        print(f'        Planilha encontrada: {sheet} / {table} - matched_sheet = {matched_sheet}')
                        break
            if matched_sheet:
                break
        
        #print(f'\n\nmatched_sheet = {matched_sheet}\n\n')

        if matched_sheet:        
            print(f"    Planilha correspondente encontrada: {sheet}")
        else:
            print(f"Nenhuma planilha encontrada que contenha '{tableNames}' no nome.")


        print(f'\nPlanilha_name = {planilha_name} | Matched_sheet = {matched_sheet} / {normalize_text(matched_sheet)} | tableNames = {tableNames}\n')
        
        if tableReferenciaName:
            print(f'    >>> PRIMEIRA | tableReferenciaName = {tableReferenciaName} - type = {type(tableReferenciaName)}')
            for i, table in enumerate(tableReferenciaName):
                table_norm = normalize_text(table)
                print(f'        table = {table_norm} | i {i} | tableReferenciaName[i] = {table}')
                if table_norm == 'ISD':
                    print(f'        table = {table}')
                    split_id = 5
                    header_id = 9
                elif table_norm == 'CSD':
                    print(f'        table = {table}')
                    split_id = 4
                    header_id = 9
                elif table_norm == 'ANALITICO':
                    print(f'        table = {table}')
                    split_id = 0
                    header_id = 9
        else:
            print(f'    >>> SEGUNDA | tableReferenciaName = {tableReferenciaName}')
            try:
                if tableNames[0] in normalize_text(matched_sheet): # Insumos Coeficiente
                    split_id = 5
                    header_id = 5
                elif tableNames[1] in normalize_text(matched_sheet): # Códigos Manutenções
                    split_id = 0
                    header_id = 5
                elif tableNames[2] in normalize_text(matched_sheet): # Mão de Obra
                    split_id = 4
                    header_id = 5
                # elif tableNames[3] in normalize_text(matched_sheet): # Composições
                #     split_id = 0
                #     header_id = 9
                #     for table in normalize_text(tableReferenciaName):
                #         if table == 'ISD':
                #             split_id = 5
                #             header_id = 9
                #         elif table == 'CSD':
                #             split_id = 4
                #             header_id = 9
                #         elif table == 'ANALITICO':
                #             split_id = 0
                #             header_id = 9                
                else:
                    print('Não foram identificadas planilhas válidas, portanto o id de coluna não foi definido')
                    split_id = 0
                    header_id = 0                
            except Exception as e:
                print(f'Filtro de colunas não encontrado - {e}')
                exit()
            print(f'header_id = {header_id}')

        
        def inserir_dados_df(file_path, matched_sheet, header_id, split_id=0):
            """
            Lê dados de uma planilha Excel, normaliza os nomes das colunas e, opcionalmente,
            realiza o 'melt' (unpivot) nos dados.

            Args:
            file_path (str): Caminho para o arquivo Excel.
            matched_sheet (str): Nome da planilha a ser lida.
            header_id (int): Linha de cabeçalho na planilha.
            split_id (int, optional): Índice da coluna para realizar o 'melt'. Se 0, não realiza o 'melt'. Defaults to 0.

            Returns:
            pd.DataFrame: DataFrame resultante.
            """
            df = pd.read_excel(file_path, sheet_name=matched_sheet, header=header_id)
            base = df.copy()
            base.columns = [normalize_text(col) for col in base.columns]

            if split_id != 0:
                print(f'    split_id = {split_id}')
                df_normalized = base.melt(
                    id_vars=base.columns[:split_id],
                    value_vars=base.columns[split_id:],
                    var_name='Estado',
                    value_name='Coeficiente'
                )
                df_normalized.columns = [normalize_text(col) for col in df_normalized.columns]
            else:
                df_normalized = base.copy()

            return df_normalized






        if isinstance(header_id, list):
            df_normalized = []
            for i,header in enumerate(header_id):
                df_normalized.append(inserir_dados_df(matched_sheet,header))
                print(df_normalized.head())
            
        else:
            df_normalized = inserir_dados_df(file_path,matched_sheet,header_id,split_id)
            print(df_normalized.head())
            

        print(type(df_normalized))

    except Exception as e:
        print(f'erro: {e}')

    except Exception as KeyboardInterrupt:
        print('\n\nFinalização do programa pelo usuário\n')
        #exit()

    


if __name__ == "__main__":
    prog = False
    while prog == False:
        print('\n===============================================================')
        try:
            main()
        except Exception as e:
            choice = input(f'\n{e}\n\nDeseja continuar o Programa (S/N)? ')
            if choice.upper() == 'S':
                prog = True
            else:
                prog = False
                break
