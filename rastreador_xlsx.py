"""
Rastreador de Arquivos Excel (.xlsx)
Este script percorre um diretório especificado pelo usuário, analisa todos os arquivos Excel (.xlsx) 
encontrados e extrai informações sobre o número de células preenchidas, 
linhas e colunas de cada planilha. Os resultados são formatados e salvos em um log JSON e TXT.
"""
import os
import json
from datetime import datetime
from openpyxl import load_workbook

def scan_excel_directory(diretorio):
    """Escaneia o diretório e retorna o resultado do processamento dos arquivos .xlsx"""
    resultado = {}
    for arquivo in os.listdir(diretorio):
        if not arquivo.endswith('.xlsx'):
            continue
        caminho = os.path.join(diretorio, arquivo)
        try:
            wb = load_workbook(caminho, read_only=True)
            planilhas_info = []
            for nome_planilha in wb.sheetnames:
                ws = wb[nome_planilha]
                dados = get_sheet_data(ws)
                planilhas_info.append((nome_planilha, dados))
            resultado[arquivo] = planilhas_info
            wb.close()
        except Exception as e:
            print(f"Erro ao processar {arquivo}: {str(e)}")
    return resultado

def get_timestamp():
    """Retorna timestamp formatado para nome de arquivo"""
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

def get_sheet_data(ws):
    """Calcula células preenchidas, linhas e colunas do retângulo mínimo de dados"""
    if not ws.cell(row=1, column=1).value and not ws.cell(row=1, column=2).value and not ws.cell(row=2, column=1).value:
        return [0, 0, 0]
    
    min_row = ws.min_row
    max_row = ws.max_row
    min_col = ws.min_column
    max_col = ws.max_column
    
    total_cells = 0
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                          min_col=min_col, max_col=max_col,
                          values_only=True):
        for cell in row:
            if cell is None:
                continue
            if isinstance(cell, str) and cell.strip() == '':
                continue
            total_cells += 1
    
    n_rows = max_row - min_row + 1
    n_cols = max_col - min_col + 1
    
    return [total_cells, n_rows, n_cols]

def format_output(result_dict):
    """Formata a saída conforme o padrão especificado"""
    output = []
    for arquivo, planilhas in result_dict.items():
        plural = 's' if len(planilhas) > 1 else ''
        output.append(f"\n    {arquivo} - {len(planilhas)} Planilha{plural}:")
        
        for planilha in planilhas:
            nome, dados = planilha
            celulas, linhas, colunas = dados
            plural_celulas = 's' if celulas != 1 else ''
            plural_linhas = 's' if linhas != 1 else ''
            plural_colunas = 's' if colunas != 1 else ''
            output.append(f"        {nome}: {celulas:,} célula{plural_celulas} ({colunas} coluna{plural_colunas} e {linhas} linha{plural_linhas})".replace(",", "."))
        
    return "\n".join(output)

def save_log(directory, data, formatted_output):
    """Salva log com dados completos em formato JSON"""
    log_data = {
        "timestamp": datetime.now().isoformat(),
        "directory": directory,
        "raw_data": data,
        "formatted_output": formatted_output
    }
    
    log_filename = f"excel_scan_log_{get_timestamp()}.json"
    log_path = os.path.join(directory, log_filename)
    
    with open(log_path, 'w', encoding='utf-8') as f:
        json.dump(log_data, f, indent=1, ensure_ascii=False)

    with open(log_path.replace('.json', '.txt'), 'w', encoding='utf-8') as f:
        f.write("=== DADOS COMPLETOS DA ANÁLISE ===\n\n")
        f.write("1. METADADOS:\n")
        f.write(f"Data/hora: {log_data['timestamp']}\n")
        f.write(f"Diretório analisado: {log_data['directory']}\n")
        f.write(f"Total de arquivos .xlsx processados: {len(data)}\n\n")
        
        f.write("2. SAÍDA FORMATADA:\n")
        f.write(formatted_output)
        f.write("\n\n=========================================\n\n")
        
    return log_path

def main():
    """Função principal para executar o script"""
    # Configuração do diretório
    diretorio = input("\nDigite o caminho do diretório (ou deixe em branco para usar o atual): ").strip()

    if not diretorio:
        diretorio = os.getcwd()
    else:
        diretorio = diretorio.replace("\\", "/").replace("'", "").strip('"')
        
    print(f'\n    Processando arquivos Excel no diretório: {diretorio}')

    # Processamento dos arquivos
    resultado = {}
    for arquivo in os.listdir(diretorio):
        if not arquivo.endswith('.xlsx'):
            continue
            
        caminho = os.path.join(diretorio, arquivo)
        try:
            wb = load_workbook(caminho, read_only=True)
            planilhas_info = []
            
            for nome_planilha in wb.sheetnames:
                ws = wb[nome_planilha]
                dados = get_sheet_data(ws)
                planilhas_info.append((nome_planilha, dados))
            
            resultado[arquivo] = planilhas_info
            wb.close()
        except Exception as e:
            print(f"Erro ao processar {arquivo}: {str(e)}")

    # Geração da saída formatada
    formatted = format_output(resultado)
    print(f"""
    ==========================================================================
    ARQUIVOS PROCESSADOS:
    {formatted}
    ==========================================================================
    """)

    # Salvando o log
    log_path = save_log(diretorio, resultado, formatted)
    print(f"\n    Log completo salvo em: {log_path}\n")

if __name__ == "__main__":
    main()

