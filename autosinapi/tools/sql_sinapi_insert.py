import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import autosinapi as sinapi
from sinapi_utils import SinapiProcessor, SinapiDownloader, SinapiLogger, DatabaseManager, FileManager

# Tenta acessar um atributo de cada classe/função importada para verificar se a importação foi bem-sucedida
try:
    _ = SinapiProcessor.process_data
    _ = SinapiDownloader.download
    _ = SinapiLogger.log
    _ = DatabaseManager.connect
    _ = FileManager.normalize_text
    print("Importação de módulos e funções de autosinapi.sinapi_utils bem-sucedida!")
except AttributeError as e:
    print(f"Erro ao verificar a importação: {e}. Algum módulo ou função pode não estar disponível.")
except ImportError as e:
    print(f"Erro de importação: {e}. Verifique se o módulo 'autosinapi' está instalado corretamente e se o caminho está correto.")


normalize_text = FileManager.normalize_text

def inserir_dados_df(file_path, matched_sheet, header_id, split_id=0):
            """
            Lê dados de uma planilha Excel, normaliza os nomes das colunas e, opcionalmente,
            realiza o 'melt' (unpivot) nos dados.

            Args:
            file_path (str): Caminho para o arquivo Excel.
            matched_sheet (str): Nome da planilha a ser lida.
            header_id (int): Linha de cabeçalho na planilha.
            split_id (int, optional): Índice da coluna para realizar o 'melt'. Se 0, não realiza o 'melt'. Defaults to 0.

            Returns:
            pd.DataFrame: DataFrame resultante.
            """
            df = pd.read_excel(file_path, sheet_name=matched_sheet, header=header_id)
            base = df.copy()
            base.columns = [normalize_text(col) for col in base.columns]

            if split_id != 0:
                print(f'    split_id = {split_id}')
                df_normalized = base.melt(
                    id_vars=base.columns[:split_id],
                    value_vars=base.columns[split_id:],
                    var_name='Estado',
                    value_name='Coeficiente'
                )
                df_normalized.columns = [normalize_text(col) for col in df_normalized.columns]
            else:
                df_normalized = base.copy()

            return df_normalized

def main():
    try:
        # perguntar o ano e mes de referencia da base sinapi para insersão no banco de dados

        data_ano_referencia = input(f'Digite o ano (YYYY): ')
        data_mes_referencia = input(f'Digite o mes (MM): ')
        formato_arquivos = input(f'Digite o formato dos arquivos (1 - xlsx, 2 - pdf): ')
        
        # Validando e Tratando mes e ano
        if formato_arquivos:
            try:
                if formato_arquivos in ['1', '2']: 
                    if formato_arquivos == '1':
                        formato = 'xlsx'
                    elif formato_arquivos == '2':
                        formato = 'pdf'
                elif formato_arquivos in ['xlsx', 'pdf']:
                    if formato_arquivos == 'xlsx':
                        formato = 'xlsx'
                    elif formato_arquivos == 'pdf':
                        formato = 'pdf'
            except ValueError:
                    print(f'Valor {formato_arquivos} inválido para o formato do arquivo')
        
        if sinapi.SinapiDownloader._validar_parametros(data_ano_referencia, data_mes_referencia, formato_arquivos) is False:
            exit()
        
        elif data_ano_referencia:
            ano = data_ano_referencia
        elif data_mes_referencia:
            mes = data_mes_referencia
        
        
        print(f'referencia: {mes}/{ano} | formato: {formato}')
        exit()                
        
        # verificando se arquivos existem na pasta e listando-os ou perguntando se o usuário quer fazer o download da base.
        diretorio_atual = os.getcwd()
        diretorio = f'{diretorio_atual}/{ano}_{mes}/SINAPI-{ano}-{mes}-formato-{formato}'
        diretorio = diretorio.replace('/', '\\')
        resultado = rastreador_xlsx.scan_excel_directory(diretorio)

        if resultado == None:
            print('Não foram encontrados arquivos no diretório: ', diretorio)
            exit()

        planilhas = list(resultado.keys())

        #print(f'planilhas: {planilhas}')
        #input do usuário para definir qual planilha utilizar:

        print(f'\nEscolha a planilha que deseja utilizar:')
        menu_planilhas = [print(f'    {i} - {planilhas[i]}') for i in range(len(planilhas))]

        planilha_index = input(f'Escolha ao índice da planilha que deseja utilizar: ')
        planilha = planilhas[int(planilha_index)]
        print(f'\nPlanilha escolhida: {planilha_index} - {planilha}\n')

        file_path = f'{diretorio}/{planilha}'
        file_path = file_path.replace('\\', '/')

        planilha_name = str(planilha).split('SINAPI_')[-1].split(f'_{ano}_{mes}.')[0]
        planilha_name = planilha_name.replace('_', ' ').replace('-', ' ').upper()

        tableNames = ('COEFICIENTES', 'MANUTENCOES', 'SEM_DESONERACAO', 'ANALITICO', 'PRECOS')

        #print(f'    Planilha-Name: {planilha_name}')

        # lista todas as planilhas do arquivo excel e verifica se o planilha_name tem relação com algum nome
        # Lista todas as planilhas do arquivo Excel
        workbook = load_workbook(file_path, read_only=True)
        all_sheets = workbook.sheetnames
        print(f"\n    Planilhas disponíveis no arquivo: {all_sheets}\n    TableNames: {tableNames}\n")

        # Verifica se planilha_name tem relação com alguma das planilhas do arquivo
        matched_sheet = [] if 'REFERENCIA' in planilha_name else None
        tableReferenciaName = [] if 'REFERENCIA' in planilha_name else None

        for sheet in all_sheets:
            normalized_sheet = normalize_text(sheet.upper())
            print(f'    Planilha: {normalized_sheet}')
            if isinstance(matched_sheet, list):
                tableReferenciaName = ['ISD','CSD','ANALITICO']

            # Se for lista, verificar se todos os itens de tableNames estão na sheet
            if all(item in normalized_sheet for item in tableReferenciaName):
                matched_sheet.append(sheet)
                print(f'        Todas as tabelas {tableReferenciaName} encontradas em: {sheet}')
            else:
                # Caso contrário, procurar por cada tabela individualmente
                for table in tableNames:
                    if table in normalized_sheet:
                        matched_sheet = sheet
                        print(f'        Planilha encontrada: {sheet} / {table} - matched_sheet = {matched_sheet}')
                        break
            if matched_sheet:
                break
        
        #print(f'\n\nmatched_sheet = {matched_sheet}\n\n')

        if matched_sheet:        
            print(f"    Planilha correspondente encontrada: {sheet}")
        else:
            print(f"Nenhuma planilha encontrada que contenha '{tableNames}' no nome.")


        print(f'\nPlanilha_name = {planilha_name} | Matched_sheet = {matched_sheet} / {normalize_text(matched_sheet)} | tableNames = {tableNames}\n')
        
        if tableReferenciaName:
            print(f'    >>> PRIMEIRA | tableReferenciaName = {tableReferenciaName} - type = {type(tableReferenciaName)}')
            for i, table in enumerate(tableReferenciaName):
                table_norm = normalize_text(table)
                print(f'        table = {table_norm} | i {i} | tableReferenciaName[i] = {table}')
                if table_norm == 'ISD':
                    print(f'        table = {table}')
                    split_id = 5
                    header_id = 9
                elif table_norm == 'CSD':
                    print(f'        table = {table}')
                    split_id = 4
                    header_id = 9
                elif table_norm == 'ANALITICO':
                    print(f'        table = {table}')
                    split_id = 0
                    header_id = 9
        else:
            print(f'    >>> SEGUNDA | tableReferenciaName = {tableReferenciaName}')
            try:
                if tableNames[0] in normalize_text(matched_sheet): # Insumos Coeficiente
                    split_id = 5
                    header_id = 5
                elif tableNames[1] in normalize_text(matched_sheet): # Códigos Manutenções
                    split_id = 0
                    header_id = 5
                elif tableNames[2] in normalize_text(matched_sheet): # Mão de Obra
                    split_id = 4
                    header_id = 5
                # elif tableNames[3] in normalize_text(matched_sheet): # Composições
                #     split_id = 0
                #     header_id = 9
                #     for table in normalize_text(tableReferenciaName):
                #         if table == 'ISD':
                #             split_id = 5
                #             header_id = 9
                #         elif table == 'CSD':
                #             split_id = 4
                #             header_id = 9
                #         elif table == 'ANALITICO':
                #             split_id = 0
                #             header_id = 9                
                else:
                    print('Não foram identificadas planilhas válidas, portanto o id de coluna não foi definido')
                    split_id = 0
                    header_id = 0                
            except Exception as e:
                print(f'Filtro de colunas não encontrado - {e}')
                exit()
            print(f'header_id = {header_id}')

        
        
        if isinstance(header_id, list):
            df_normalized = []
            for i,header in enumerate(header_id):
                df_normalized.append(inserir_dados_df(matched_sheet,header))
                print(df_normalized.head())
            
        else:
            df_normalized = inserir_dados_df(file_path,matched_sheet,header_id,split_id)
            print(df_normalized.head())
            

        #print(type(df_normalized))

    except Exception as e:
        print(f'erro: {e}')

    except Exception as KeyboardInterrupt:
        print('\n\nFinalização do programa pelo usuário\n')
        #exit()

if __name__ == "__main__":
    prog = False
    while prog == False:
        print('\n===============================================================')
        try:
            main()
        except Exception as e:
            choice = input(f'\n{e}\n\nDeseja continuar o Programa (S/N)? ')
            if choice.upper() == 'S':
                prog = True
            else:
                prog = False
                break
